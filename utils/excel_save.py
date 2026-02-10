# app/exporters/validation_cart_xlsx.py
from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook


def export_cart_xlsx(
    *,
    columns: Sequence[Mapping[str, Any]],                               # Sequence[dict]
    out_path: Path = Path("data/validation_cart.xlsx"),
    start_col: int = 1,   # A열부터
) -> Path:
    """
    1~3행만 사용해서 '컬럼 블록'을 가로로 쭉 작성한다.

    columns: [
      {"title": "...", "items": "...", "values": "..."},
      ...
    ]
    title: 1행에 들어갈 제목
    items: 2행에 들어갈 적용 문항(쉼표로 구분)
    values: 3행에 들어갈 규칙 값 (범위 등)
    """

    wb = Workbook()     # 새 워크북 생성
    ws = wb.active      # 활성 워크시트 선택

    for i, c in enumerate(columns):
        col = start_col + i
        ws.cell(row=1, column=col, value=c.get("title", ""))    # 해당 열과 행에 값 설정(title값, 없으면 빈문자열)
        ws.cell(row=2, column=col, value=c.get("items", ""))
        ws.cell(row=3, column=col, value=c.get("values", ""))

    wb.save(out_path)
    return out_path