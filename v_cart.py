# app/exporters/validation_cart_xlsx.py
from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook


def export_cart_xlsx(
    *,
    columns: Sequence[Mapping[str, Any]],
    out_path: str | Path = "data/validation_cart.xlsx",
    start_col: int = 1,   # A열부터
) -> Path:
    """
    1~3행만 사용해서 '컬럼 블록'을 가로로 쭉 작성한다.

    columns: [
      {"title": "...", "items": "...", "values": "..."},
      ...
    ]
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active

    for i, c in enumerate(columns):
        col = start_col + i
        ws.cell(row=1, column=col, value=c.get("title", ""))
        ws.cell(row=2, column=col, value=c.get("items", ""))
        ws.cell(row=3, column=col, value=c.get("values", ""))

    wb.save(out_path)
    return out_path


print("v_cart loaded, functions:", dir())
