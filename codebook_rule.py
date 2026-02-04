import pandas as pd
from openpyxl import Workbook


# 파일 불러오기 및 전처리
cb = pd.read_excel("data/test.xlsx", sheet_name="codebook")

codebook_df = cb[["문항", "응답"]].dropna(subset=["문항"]).reset_index(drop=True)

# range_items 범위 지정
def extract_codes(응답: str) -> list[int]:
    """
    코드북의 '응답' 문자열에서 선택지 코드를 추출한다.

    매개변수:
        응답 (str): 코드북의 '응답' 컬럼 문자열.
                   예: "1: 남자\\n2: 여자\\n3: 기타"

    반환값:
        list[int]:
            - 형식이 올바른 경우: 추출된 정수 코드들의 리스트   예: [1, 2, 3]
            - 형식이 올바르지 않은 경우: 빈 리스트 []          (범위 검증 대상 아님을 의미)
            
    내부 변수
        line (str): 줄 단위로 분리된 응답 문자열의 한 행.
        colon_index (int): line 문자열에서 첫 번째 ':' 문자가 등장하는 위치(index).
        code_str (str): line에서 ':' 이전 부분을 잘라낸 문자열.
        code_values (list[int]): 형식 검증을 통과한 정수 코드들을 누적 저장하는 리스트.
    """
    code_values = []
    for raw in 응답.splitlines():  # 문자열을 줄단위로 쪼개기 ex)["1: 남자", "2: 여자", "3: 기타"]
        line = raw.strip()  # 앞뒤 공백 제거
        if not line:
            continue

        # 1) 구분자 찾기: ':' 우선, 없으면 '.'
        colon_index = line.find(":")
        if colon_index == -1:     # -1의 의미 : 해당 문자를 찾지못함
            return []

        # 2) 구분자 앞 부분을 코드 후보로
        code_str = line[:colon_index].strip()

        # 3) 숫자인 경우만 코드로 인정
        if code_str.isdigit():              # isdigit() : 문자열이 숫자로만 이루어져있는지 확인
            code_values.append(int(code_str))

    return code_values

# 'codes_values' 컬럼에 추출된 코드 리스트 저장
codebook_df["codes_values"] = codebook_df["응답"].apply(extract_codes)

codebook_df["check_range"] = codebook_df["codes_values"].apply(bool)

codebook_df["range_min"] = codebook_df["codes_values"].apply(
    lambda x: min(x) if x else None
)
codebook_df["range_max"] = codebook_df["codes_values"].apply(
    lambda x: max(x) if x else None
)

codebook_df["check_missing"] = True
codebook_df["check_multi"] = True


# 검증 항목별 문항 리스트 생성
missing_items = codebook_df.loc[codebook_df["check_missing"], "문항"].tolist()
multiple_items = codebook_df.loc[codebook_df["check_multi"], "문항"].tolist()
range_items = codebook_df.loc[codebook_df["check_range"], "문항"].tolist()






print(codebook_df[["문항", "codes_values", "check_range", "range_min", "range_max"]].head())
print(
    codebook_df.loc[
        codebook_df["문항"] == "B2",
        ["문항", "codes_values", "check_range", "range_min", "range_max"]
    ]
)

# 검증 요약 데이터프레임 생성
summary_df = pd.DataFrame({
    "결측": [",".join(missing_items)],
    "중복응답": [",".join(multiple_items)],
    "범위": [",".join(range_items)]
})



#########################

# 1) range 그룹 만들기 (min,max별로 문항 묶기)
g = (codebook_df.loc[codebook_df["check_range"], ["문항", "range_min", "range_max"]]
    .groupby(["range_min", "range_max"])["문항"]
    .apply(lambda s: ", ".join(s))
    .reset_index(name="items"))

# 2) 엑셀 작성
out_path = "data/validation_cart.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "검증_장바구니"

# 결측 / 중복응답 (한 셀에 쭉)
ws["A1"] = "결측"
ws["A2"] = ", ".join(missing_items)

ws["B1"] = "중복응답"
ws["B2"] = ", ".join(multiple_items)

# 범위 블록 (C열부터)
start_col = 3  # C
for i, r in enumerate(g.itertuples(index=False), start=0):
    col = start_col + i
    ws.cell(row=1, column=col, value="범위")
    ws.cell(row=2, column=col, value=r.items)
    ws.cell(row=3, column=col, value=f"{int(r.range_min)}, {int(r.range_max)}")

wb.save(out_path)
print(f"saved: {out_path}")

